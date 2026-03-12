"""
FTA Risk Allocator v10
Clean rewrite — robust allocation engine, shared failures, persistent save/load.

Core rules:
  OR  gate top-down:  child_T = parent_T / n
  AND gate top-down:  child_T = parent_T ^ (1/n)
  OR  gate rollup:    parent_A = sum(children_A)
  AND gate rollup:    parent_A = prod(children_A)

Allocation and Achieved are COMPLETELY SEPARATE pipelines.
"""

import streamlit as st
import streamlit.components.v1 as components
import json, math, io, datetime, os, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── page config ────────────────────────────────────────────────
st.set_page_config(page_title="FTA Allocator v10", page_icon="⚛", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=DM+Sans:wght@300;400;500;600&display=swap');
*{box-sizing:border-box}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif}
.stApp{background:#080c14;color:#d4dde8}
section[data-testid="stSidebar"]{background:#0c1220!important;border-right:1px solid #1e2d45}
section[data-testid="stSidebar"] *{color:#d4dde8!important}
.stButton>button{background:#0f1d30!important;border:1px solid #1e3a5f!important;
  color:#7ab8e8!important;border-radius:5px!important;font-family:'DM Sans',sans-serif!important;
  font-size:0.8rem!important;transition:all .15s!important}
.stButton>button:hover{background:#1e3a5f!important;color:#b8d8f5!important;border-color:#4a8cc2!important}
.stTabs [data-baseweb="tab-list"]{background:#0c1220;border-bottom:1px solid #1e2d45}
.stTabs [data-baseweb="tab"]{color:#5a7a9a!important;font-family:'DM Sans',sans-serif}
.stTabs [aria-selected="true"]{color:#7ab8e8!important;border-bottom:2px solid #4a8cc2!important}
div[data-testid="stExpander"]{background:#0c1220;border:1px solid #1e2d45;border-radius:6px}
.stTextInput input,.stNumberInput input,.stSelectbox select,.stTextArea textarea{
  background:#0c1220!important;border:1px solid #1e2d45!important;color:#d4dde8!important;
  font-family:'JetBrains Mono',monospace!important;border-radius:5px!important}
.stSelectbox [data-baseweb="select"]>div{background:#0c1220!important;border-color:#1e2d45!important}
hr{border-color:#1e2d45!important}

/* Value badges */
.vbadge{font-family:'JetBrains Mono',monospace;font-size:0.78rem;font-weight:600;
  display:inline-block;padding:2px 9px;border-radius:4px;letter-spacing:.5px}
.vb-alloc{background:#0d2036;border:1px solid #1e4a7a;color:#5aabff}
.vb-ach{background:#0d1f18;border:1px solid #1e4a32;color:#4ade80}
.vb-ach.over{background:#200d0d;border-color:#4a1e1e;color:#f87171}
.vb-none{background:#141820;border:1px solid #2a3040;color:#4a5a6a}

/* Node type badges */
.nb{display:inline-block;padding:1px 7px;border-radius:3px;font-size:0.67rem;font-weight:700;
  font-family:'JetBrains Mono',monospace;letter-spacing:.5px}
.nb-HZ{background:#2d1200;color:#fb923c;border:1px solid #7c3300}
.nb-SF{background:#001830;color:#60a5fa;border:1px solid #1e4878}
.nb-FF{background:#001a10;color:#34d399;border:1px solid #065f35}
.nb-IF{background:#1a0030;color:#c084fc;border:1px solid #5b21b6}
.nb-AND{background:#200025;color:#e879f9;border:1px solid #7e22ce}

/* Gate badge */
.gb{display:inline-block;padding:1px 6px;border-radius:3px;font-size:0.65rem;font-weight:700;
  font-family:'JetBrains Mono',monospace}
.gb-OR{background:#001828;color:#38bdf8;border:1px solid #0369a1}
.gb-AND{background:#1e0028;color:#d946ef;border:1px solid #86198f}

/* Shared tag */
.st-shared{display:inline-block;padding:1px 6px;border-radius:3px;font-size:0.62rem;
  font-weight:700;background:#1a1000;color:#fbbf24;border:1px solid #78350f;margin-left:4px}

/* Row styling */
.alloc-row{border-bottom:1px solid #0f1a28;padding:6px 0;transition:background .1s}
.alloc-row:hover{background:#0c1628}

/* Info callout */
.callout{background:#0c1628;border:1px solid #1e3a5f;border-left:3px solid #4a8cc2;
  border-radius:6px;padding:10px 14px;font-size:0.8rem;color:#7ab8e8;line-height:1.6;margin:8px 0}
.callout.warn{border-left-color:#f59e0b;color:#fcd34d;background:#0f1208}
.callout.ok{border-left-color:#22c55e;color:#86efac;background:#081208}

/* Header */
.app-header{background:linear-gradient(135deg,#0c1e38 0%,#080c14 100%);
  border:1px solid #1e3a5f;border-left:4px solid #4a8cc2;border-radius:8px;
  padding:16px 24px;margin-bottom:18px}
.app-header h1{font-family:'JetBrains Mono',monospace;font-size:1.3rem;
  color:#7ab8e8;margin:0 0 3px;letter-spacing:-0.5px}
.app-header p{color:#4a6a8a;margin:0;font-size:0.75rem}

/* Save status */
.save-ok{background:#081208;border:1px solid #1e4032;border-radius:5px;
  padding:5px 12px;font-size:0.72rem;color:#4ade80}
.save-no{background:#0f1208;border:1px solid #2a3010;border-radius:5px;
  padding:5px 12px;font-size:0.72rem;color:#a3a380}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════
VALID_PARENTS = {
    "SF":  ["HZ", "SF", "AND"],
    "FF":  ["SF", "FF", "AND"],
    "IF":  ["FF", "SF"],
    "AND": ["HZ", "SF", "FF"],
}
TYPE_STR   = {"HZ":"#fb923c","SF":"#60a5fa","FF":"#34d399","IF":"#c084fc","AND":"#e879f9"}
SAVE_FILE  = "fta_save.json"

# ══════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════
def _init():
    if "nodes" not in st.session_state:
        st.session_state.nodes = {}          # id -> node dict
    if "hz_targets" not in st.session_state:
        st.session_state.hz_targets = {}     # hz_id -> float
    if "nxt" not in st.session_state:
        st.session_state.nxt = 1             # id counter

_init()

# ══════════════════════════════════════════════════════════════════════
# NODE HELPERS
# ══════════════════════════════════════════════════════════════════════
def children_of(nodes, pid):
    return [n for n in nodes.values() if n.get("parent") == pid]

def hz_roots(nodes):
    return [n for n in nodes.values() if n["type"] == "HZ"]

def descendants(nodes, nid):
    result = []
    for n in nodes.values():
        if n.get("parent") == nid:
            result.append(n["id"])
            result.extend(descendants(nodes, n["id"]))
    return result

def ancestor_hz(nodes, nid):
    cur, seen = nid, set()
    while cur and cur not in seen:
        seen.add(cur)
        n = nodes.get(cur)
        if not n: return None
        if n["type"] == "HZ": return cur
        cur = n.get("parent")
    return None

def depth_of(nodes, nid):
    d, cur, seen = 0, nid, set()
    while nodes.get(cur, {}).get("parent"):
        cur = nodes[cur]["parent"]
        if cur in seen: break
        seen.add(cur); d += 1
    return d

def bfs_order(nodes):
    roots = [n["id"] for n in hz_roots(nodes)]
    out, queue, seen = [], list(roots), set()
    while queue:
        nid = queue.pop(0)
        if nid in seen: continue
        seen.add(nid); out.append(nid)
        queue.extend(c["id"] for c in children_of(nodes, nid))
    return out

def nodes_with_label(nodes, label):
    return [nid for nid, n in nodes.items() if n.get("label","") == label and label]

def fmt(v, dash="–"):
    if v is None: return dash
    if v == 0:   return "0.000E+00"
    return f"{v:.3E}"

def next_id():
    nid = f"N{st.session_state.nxt:04d}"
    st.session_state.nxt += 1
    return nid

# ══════════════════════════════════════════════════════════════════════
# CORE ENGINE 1: TOP-DOWN ALLOCATION
# Rule: allocation is PURE MATHS — never reads achieved values
# ══════════════════════════════════════════════════════════════════════
def allocate(nodes, hz_targets):
    """
    Recursively assign allocated budget to every node.
    Returns dict {nid: float}.
    OR  gate: each child = parent / n_children
    AND gate: each child = parent ^ (1/n_children)
    """
    result = {}

    def _recurse(nid, budget):
        result[nid] = budget
        kids = children_of(nodes, nid)
        if not kids:
            return
        n = len(kids)
        for kid in kids:
            gate = kid.get("gate", "OR")
            if gate == "AND":
                # AND: each child gets parent^(1/n) — combined prob = product
                child_budget = (budget ** (1.0 / n)) if budget > 0 else 0.0
            else:
                # OR: each child gets parent/n — combined prob = sum
                child_budget = budget / n if n > 0 else 0.0
            _recurse(kid["id"], child_budget)

    for hz in hz_roots(nodes):
        target = hz_targets.get(hz["id"], 1e-7)
        _recurse(hz["id"], target)

    return result

# ══════════════════════════════════════════════════════════════════════
# CORE ENGINE 2: BOTTOM-UP ROLLUP OF ACHIEVED VALUES
# Rule: rollup is PURE MATHS — never reads alloc
# ══════════════════════════════════════════════════════════════════════
def rollup(nodes):
    """
    Compute achieved (rolled-up) value for every node.
    Leaf nodes: use node["achieved"] if set, else None.
    Non-leaf: use gate logic over children. If any child is None → None.
    Manual override: if node has achieved set AND has children, use manual value
    (allows partial tree entry).
    Returns dict {nid: float|None}.
    """
    cache = {}

    def _compute(nid):
        if nid in cache:
            return cache[nid]
        n = nodes.get(nid)
        if n is None:
            cache[nid] = None; return None

        kids = children_of(nodes, nid)

        # Leaf node — use manually entered value
        if not kids:
            cache[nid] = n.get("achieved")
            return cache[nid]

        # Non-leaf: compute from children first
        child_vals = [_compute(k["id"]) for k in kids]

        # If manual override exists on this node, use it
        manual = n.get("achieved")
        if manual is not None:
            cache[nid] = manual
            return manual

        # If any child has no value → can't roll up
        if any(v is None for v in child_vals):
            cache[nid] = None
            return None

        gate = n.get("gate", "OR")
        if gate == "AND":
            val = 1.0
            for v in child_vals:
                val *= v
        else:  # OR
            val = sum(child_vals)

        cache[nid] = val
        return val

    for hz in hz_roots(nodes):
        _compute(hz["id"])
    # Catch any orphaned nodes
    for nid in nodes:
        if nid not in cache:
            _compute(nid)
    return cache

# ══════════════════════════════════════════════════════════════════════
# CORE ENGINE 3: SHARED FAILURE SYNC
# When achieved on node N changes → all nodes with same label get
# worst-case (max) achieved value.
# Does NOT touch allocation.
# ══════════════════════════════════════════════════════════════════════
def sync_shared(nodes, changed_nid, new_value):
    """
    Propagate worst-case achieved value to all nodes sharing the same label.
    Returns list of (nid, old_val, new_val) for change log.
    """
    label = nodes[changed_nid].get("label", "")
    peers = nodes_with_label(nodes, label)
    if len(peers) <= 1:
        # Only node with this label — just set it
        old = nodes[changed_nid].get("achieved")
        nodes[changed_nid]["achieved"] = new_value
        return [(changed_nid, old, new_value)] if old != new_value else []

    # Worst-case: max of new value and all existing peer achieved values
    existing = [nodes[p].get("achieved") for p in peers if nodes[p].get("achieved") is not None]
    worst = max([new_value] + existing)

    log = []
    for pid in peers:
        old = nodes[pid].get("achieved")
        if old != worst:
            nodes[pid]["achieved"] = worst
            log.append((pid, old, worst))
    return log

# ══════════════════════════════════════════════════════════════════════
# PERSISTENCE: JSON FILE + DOWNLOAD/UPLOAD
# ══════════════════════════════════════════════════════════════════════
def state_to_dict():
    return {
        "version": "v10",
        "saved_at": datetime.datetime.utcnow().isoformat() + "Z",
        "nodes": st.session_state.nodes,
        "hz_targets": st.session_state.hz_targets,
        "nxt": st.session_state.nxt,
    }

def dict_to_state(d):
    st.session_state.nodes      = d.get("nodes", {})
    st.session_state.hz_targets = d.get("hz_targets", {})
    st.session_state.nxt        = d.get("nxt", 1)

def save_to_file():
    """Save to local filesystem (works on Streamlit Cloud via /tmp)."""
    try:
        with open(SAVE_FILE, "w") as f:
            json.dump(state_to_dict(), f, indent=2)
        return True
    except Exception:
        return False

def load_from_file():
    """Load from local filesystem if exists."""
    try:
        if os.path.exists(SAVE_FILE):
            with open(SAVE_FILE) as f:
                d = json.load(f)
            dict_to_state(d)
            return d.get("saved_at","")
    except Exception:
        pass
    return None

# Auto-load on first run
if "auto_loaded" not in st.session_state:
    st.session_state.auto_loaded = True
    ts = load_from_file()
    if ts:
        st.session_state["_last_saved"] = ts

# Auto-save every render (if data exists)
if st.session_state.nodes:
    save_to_file()
    st.session_state["_last_saved"] = datetime.datetime.utcnow().strftime("%H:%M:%S UTC")

# ══════════════════════════════════════════════════════════════════════
# VISUALIZATION — Canvas-based tree (handles 500+ nodes via BFS layout)
# ══════════════════════════════════════════════════════════════════════
def build_canvas(nodes, alloc, rolled):
    """Build HTML/JS canvas visualization. Efficient for large trees."""
    order = bfs_order(nodes)
    hz_ids = [n["id"] for n in hz_roots(nodes)]
    palette = ["#4a8cc2","#22c55e","#f59e0b","#e879f9","#38bdf8","#f87171","#a3e635","#06b6d4"]
    hz_color = {hid: palette[i % len(palette)] for i, hid in enumerate(hz_ids)}

    node_data, edge_data = [], []
    for nid in order:
        if nid not in nodes: continue
        n = nodes[nid]
        t_val = alloc.get(nid)
        a_val = rolled.get(nid)
        status = "pass" if (t_val is not None and a_val is not None and a_val <= t_val) else \
                 "fail" if (t_val is not None and a_val is not None and a_val > t_val) else "na"
        hz_anc = ancestor_hz(nodes, nid)
        peers  = nodes_with_label(nodes, n.get("label",""))
        node_data.append({
            "id": nid, "label": n.get("label", nid), "name": n.get("name",""),
            "type": n["type"], "gate": n.get("gate","–"),
            "T": fmt(t_val), "A": fmt(a_val),
            "T_raw": t_val, "A_raw": a_val,
            "status": status, "color": hz_color.get(hz_anc, "#4a8cc2"),
            "parent": n.get("parent",""),
            "shared": len(peers) > 1,
            "achieved_set": n.get("achieved") is not None,
        })
        if n.get("parent") and n["parent"] in nodes:
            edge_data.append({"from": n["parent"], "to": nid, "gate": n.get("gate","OR")})

    nj = json.dumps(node_data)
    ej = json.dumps(edge_data)

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#080c14;font-family:'JetBrains Mono',monospace;overflow:hidden;user-select:none}}
#cv{{cursor:grab;display:block}}#cv.pan{{cursor:grabbing}}
#hud{{position:absolute;top:10px;left:10px;display:flex;flex-direction:column;gap:4px;z-index:10}}
.hbtn{{background:#0c1628;border:1px solid #1e3a5f;color:#7ab8e8;padding:4px 10px;border-radius:4px;
  cursor:pointer;font-size:10px;font-family:inherit;white-space:nowrap}}
.hbtn:hover{{background:#1e3a5f;color:#b8d8f5}}
#search{{position:absolute;top:10px;left:50%;transform:translateX(-50%);
  background:#0c1628;border:1px solid #1e3a5f;color:#d4dde8;padding:5px 12px;
  border-radius:4px;font-size:11px;font-family:inherit;width:220px;outline:none}}
#search:focus{{border-color:#4a8cc2}}
#search::placeholder{{color:#3a5a7a}}
#tip{{position:absolute;background:#0c1e38;border:1px solid #1e3a5f;border-radius:6px;
  padding:8px 12px;font-size:10px;color:#d4dde8;z-index:20;display:none;
  max-width:260px;pointer-events:none;line-height:1.5;box-shadow:0 8px 24px rgba(0,0,0,.8)}}
#legend{{position:absolute;bottom:8px;left:10px;font-size:9px;color:#3a5a7a;
  display:flex;gap:10px;flex-wrap:wrap}}
.ldot{{display:inline-block;width:7px;height:7px;border-radius:50%;margin-right:3px;vertical-align:middle}}
#info{{position:absolute;bottom:8px;right:10px;font-size:9px;color:#2a4a6a}}
</style></head><body>
<canvas id="cv"></canvas>
<div id="hud">
  <button class="hbtn" onclick="layout()">⊞ Layout</button>
  <button class="hbtn" onclick="zoomIn()">＋</button>
  <button class="hbtn" onclick="zoomOut()">－</button>
  <button class="hbtn" onclick="resetView()">⌖ Reset</button>
  <button class="hbtn" id="simBtn" onclick="toggleSim()">⟳ Physics OFF</button>
</div>
<input id="search" type="text" placeholder="Search…" oninput="onSearch(this.value)">
<div id="tip"></div>
<div id="legend">
  <span><span class="ldot" style="background:#22c55e"></span>A≤T (OK)</span>
  <span><span class="ldot" style="background:#f87171"></span>A>T (Exceeds)</span>
  <span><span class="ldot" style="background:#3a5a7a"></span>No data</span>
  <span><span class="ldot" style="background:#fbbf24;border-radius:2px"></span>Shared failure</span>
</div>
<div id="info">drag · scroll=zoom · hover=info</div>
<script>
const NODES = {nj};
const EDGES = {ej};
const BW=160, BH=64, GR=11;
const STATUS_COL = {{pass:"#22c55e", fail:"#f87171", na:"#2a3a4a"}};
const TYPE_FILL  = {{HZ:"#1a0e00",SF:"#00101e",FF:"#00110a",IF:"#0e0018",AND:"#130018"}};
const TYPE_STR   = {{HZ:"#fb923c",SF:"#60a5fa",FF:"#34d399",IF:"#c084fc",AND:"#e879f9"}};
const cv = document.getElementById('cv');
const ctx = cv.getContext('2d');
let scale=1, panX=0, panY=60, sim=false;
let drag=null, dragOff={{x:0,y:0}}, isPan=false, lastP={{x:0,y:0}}, moved=false;
let pos={{}}, searchHL=new Set(), collapsed=new Set();

function resize(){{cv.width=window.innerWidth;cv.height=window.innerHeight;}}
window.addEventListener('resize',()=>{{resize();draw();}});
resize();

// ── Layout: BFS level-based placement ──────────────────────────
function layout(){{
  pos={{}};
  const levels={{}}, hzList=[...new Set(NODES.map(n=>n.id).filter(id=>NODES.find(n=>n.id===id)?.type==='HZ'))];
  NODES.forEach(n=>{{
    let d=0,cur=n.id,seen=new Set();
    while(true){{
      const nd=NODES.find(x=>x.id===cur);if(!nd||!nd.parent||seen.has(cur))break;
      seen.add(cur);cur=nd.parent;d++;
    }}
    (levels[d]||(levels[d]=[])).push(n.id);
  }});
  // Group by HZ within each level
  Object.entries(levels).forEach(([lvl,ids])=>{{
    const byHz={{}};
    ids.forEach(id=>{{
      const n=NODES.find(x=>x.id===id);
      const hz=n?.color||'#4a8cc2';
      (byHz[hz]||(byHz[hz]=[])).push(id);
    }});
    const hzGroups=Object.values(byHz);
    let xOff=0;
    hzGroups.forEach((group,gi)=>{{
      const groupW=group.length*(BW+20)-20;
      group.forEach((id,i)=>{{
        pos[id]={{x:xOff+i*(BW+20)+BW/2, y:parseInt(lvl)*(BH+80)+80, vx:0, vy:0}};
      }});
      xOff+=groupW+60;
    }});
  }});
  // Center horizontally
  const xs=Object.values(pos).map(p=>p.x);
  const mid=(Math.min(...xs)+Math.max(...xs))/2;
  const cx2=cv.width/2/scale;
  Object.values(pos).forEach(p=>{{p.x+=cx2-mid;}});
  panX=0;panY=60;scale=1;
}}
layout();

// ── Physics sim ─────────────────────────────────────────────────
function simulate(){{
  if(!sim)return;
  const ids=Object.keys(pos);
  for(let i=0;i<ids.length;i++){{
    for(let j=i+1;j<ids.length;j++){{
      const a=pos[ids[i]],b=pos[ids[j]];
      const dx=b.x-a.x,dy=b.y-a.y,d=Math.sqrt(dx*dx+dy*dy)||1;
      const f=6000/(d*d);
      a.vx-=dx/d*f;a.vy-=dy/d*f;b.vx+=dx/d*f;b.vy+=dy/d*f;
    }}
  }}
  EDGES.forEach(e=>{{
    const a=pos[e.from],b=pos[e.to];if(!a||!b)return;
    const dx=b.x-a.x,dy=b.y-a.y,d=Math.sqrt(dx*dx+dy*dy)||1;
    const f=(d-160)*0.06;a.vx+=dx/d*f;a.vy+=dy/d*f;b.vx-=dx/d*f;b.vy-=dy/d*f;
  }});
  ids.forEach(id=>{{
    const n=NODES.find(x=>x.id===id);if(!n)return;
    pos[id].vy+=n.type==='HZ'?0:0.02*(Object.keys(pos).indexOf(id)*0.01);
    if(id!==drag){{pos[id].x+=pos[id].vx;pos[id].y+=pos[id].vy;}}
    pos[id].vx*=0.78;pos[id].vy*=0.78;
  }});
}}

// ── Helpers ───────────────────────────────────────────────────
function toW(sx,sy){{return{{x:(sx-panX)/scale,y:(sy-panY)/scale}};}}
function nodeAt(wx,wy){{
  for(let i=NODES.length-1;i>=0;i--){{
    const n=NODES[i],p=pos[n.id];if(!p)continue;
    if(wx>=p.x-BW/2&&wx<=p.x+BW/2&&wy>=p.y-BH/2&&wy<=p.y+BH/2)return n;
  }}return null;
}}
function isVisible(nid){{
  let cur=nid,seen=new Set();
  while(true){{
    const n=NODES.find(x=>x.id===cur);if(!n||!n.parent)return true;
    if(collapsed.has(n.parent))return false;
    if(seen.has(cur))return true;
    seen.add(cur);cur=n.parent;
  }}
}}
function rrect(x,y,w,h,r){{
  ctx.beginPath();ctx.moveTo(x+r,y);ctx.lineTo(x+w-r,y);ctx.arcTo(x+w,y,x+w,y+r,r);
  ctx.lineTo(x+w,y+h-r);ctx.arcTo(x+w,y+h,x+w-r,y+h,r);
  ctx.lineTo(x+r,y+h);ctx.arcTo(x,y+h,x,y+h-r,r);ctx.lineTo(x,y+r);ctx.arcTo(x,y,x+r,y,r);
  ctx.closePath();
}}

// ── Draw ──────────────────────────────────────────────────────
function draw(){{
  ctx.clearRect(0,0,cv.width,cv.height);
  ctx.save();ctx.translate(panX,panY);ctx.scale(scale,scale);

  // Edges
  EDGES.forEach(e=>{{
    if(!pos[e.from]||!pos[e.to])return;
    if(!isVisible(e.from)||!isVisible(e.to))return;
    const a=pos[e.from],b=pos[e.to];
    const gc=e.gate==='AND'?'#7e22ce':'#0369a1';
    ctx.save();
    ctx.beginPath();
    ctx.moveTo(a.x,a.y+BH/2);
    ctx.bezierCurveTo(a.x,a.y+BH/2+30,b.x,b.y-BH/2-30,b.x,b.y-BH/2);
    ctx.strokeStyle=gc;ctx.lineWidth=1.2;ctx.globalAlpha=0.6;ctx.stroke();
    // Gate circle
    const mx=(a.x+b.x)/2,my=(a.y+BH/2+b.y-BH/2)/2;
    ctx.globalAlpha=1;
    ctx.beginPath();ctx.arc(mx,my,GR,0,Math.PI*2);
    ctx.fillStyle=e.gate==='AND'?'#130018':'#00101e';ctx.fill();
    ctx.strokeStyle=gc;ctx.lineWidth=1;ctx.stroke();
    ctx.fillStyle=e.gate==='AND'?'#e879f9':'#38bdf8';
    ctx.font='bold 7px JetBrains Mono, monospace';
    ctx.textAlign='center';ctx.textBaseline='middle';
    ctx.fillText(e.gate,mx,my);
    ctx.restore();
  }});

  // Nodes
  NODES.forEach(n=>{{
    if(!pos[n.id]||!isVisible(n.id))return;
    const p=pos[n.id],x=p.x-BW/2,y=p.y-BH/2;
    const sc=STATUS_COL[n.status]||STATUS_COL.na;
    const tf=TYPE_FILL[n.type]||'#0c1628';
    const ts=TYPE_STR[n.type]||'#7ab8e8';
    const hl=searchHL.size===0||searchHL.has(n.id);

    ctx.save();ctx.globalAlpha=hl?1:0.18;

    // Shadow / glow for status
    if(n.status!=='na'&&hl){{ctx.shadowColor=sc;ctx.shadowBlur=10;}}

    // Box fill
    ctx.fillStyle=tf;rrect(x,y,BW,BH,7);ctx.fill();

    // Shared failure highlight (amber border overlay)
    if(n.shared&&hl){{
      ctx.strokeStyle='#f59e0b';ctx.lineWidth=1.5;
      ctx.setLineDash([3,3]);rrect(x-2,y-2,BW+4,BH+4,9);ctx.stroke();ctx.setLineDash([]);
    }}

    // Border
    ctx.shadowBlur=0;
    ctx.strokeStyle=sc;ctx.lineWidth=hl?1.8:1;
    rrect(x,y,BW,BH,7);ctx.stroke();

    // Type header stripe
    ctx.fillStyle=ts+'22';ctx.fillRect(x,y,BW,16);ctx.fillRect(x,y,BW,7);

    // Type label
    ctx.fillStyle=ts;ctx.font='bold 7px JetBrains Mono,monospace';
    ctx.textAlign='center';ctx.textBaseline='top';
    ctx.fillText(n.type+(n.gate!=='–'?' · '+n.gate:''),p.x,y+2);

    // Node label
    ctx.fillStyle=ts;ctx.font='bold 11px JetBrains Mono,monospace';
    ctx.textAlign='center';ctx.textBaseline='middle';
    ctx.fillText(n.label.substring(0,14),p.x,p.y-7);

    // Name (smaller)
    ctx.fillStyle='#5a7a9a';ctx.font='9px DM Sans,sans-serif';
    const nm=n.name.length>20?n.name.substring(0,19)+'…':n.name;
    ctx.fillText(nm,p.x,p.y+5);

    // T and A values
    ctx.font='7px JetBrains Mono,monospace';
    ctx.textAlign='left';ctx.fillStyle='#4a8cc2';ctx.fillText('T:'+n.T,x+5,y+BH-10);
    ctx.textAlign='right';
    ctx.fillStyle=n.status==='pass'?'#22c55e':n.status==='fail'?'#f87171':'#3a5a7a';
    ctx.fillText('A:'+n.A,x+BW-5,y+BH-10);

    // Collapse button for non-leaves
    const hasKids=EDGES.some(e=>e.from===n.id);
    if(hasKids){{
      const bx=p.x+BW/2-12,by=y+4;
      ctx.fillStyle=collapsed.has(n.id)?ts:'#1e2d45';
      ctx.beginPath();ctx.arc(bx,by+5,6,0,Math.PI*2);ctx.fill();
      ctx.strokeStyle=ts;ctx.lineWidth=.8;ctx.stroke();
      ctx.fillStyle=collapsed.has(n.id)?'#080c14':ts;
      ctx.font='bold 7px monospace';ctx.textAlign='center';ctx.textBaseline='middle';
      ctx.fillText(collapsed.has(n.id)?'▶':'▼',bx,by+5);
    }}

    ctx.restore();
  }});

  ctx.restore();
}}

// ── Events ──────────────────────────────────────────────────
cv.addEventListener('mousedown',ev=>{{
  const r=cv.getBoundingClientRect();
  const{{x:wx,y:wy}}=toW(ev.clientX-r.left,ev.clientY-r.top);
  const n=nodeAt(wx,wy);
  if(n&&isVisible(n.id)){{
    // Check collapse button
    const p=pos[n.id];const bx=p.x+BW/2-12,by=p.y-BH/2+9;
    if(EDGES.some(e=>e.from===n.id)&&Math.hypot(wx-bx,wy-by)<8){{
      collapsed.has(n.id)?collapsed.delete(n.id):collapsed.add(n.id);
      return;
    }}
    drag=n.id;dragOff={{x:wx-p.x,y:wy-p.y}};moved=false;
    cv.classList.add('pan');
  }} else {{
    isPan=true;lastP={{x:ev.clientX,y:ev.clientY}};cv.classList.add('pan');
  }}
  ev.preventDefault();
}});
window.addEventListener('mousemove',ev=>{{
  if(drag){{
    const r=cv.getBoundingClientRect();
    const{{x:wx,y:wy}}=toW(ev.clientX-r.left,ev.clientY-r.top);
    pos[drag].x=wx-dragOff.x;pos[drag].y=wy-dragOff.y;moved=true;
  }} else if(isPan){{
    panX+=ev.clientX-lastP.x;panY+=ev.clientY-lastP.y;
    lastP={{x:ev.clientX,y:ev.clientY}};
  }}
  // Tooltip
  const r=cv.getBoundingClientRect();
  const{{x:wx2,y:wy2}}=toW(ev.clientX-r.left,ev.clientY-r.top);
  const hn=nodeAt(wx2,wy2);
  const tip=document.getElementById('tip');
  if(hn&&isVisible(hn.id)){{
    const sc=hn.status==='pass'?'#22c55e':hn.status==='fail'?'#f87171':'#7ab8e8';
    tip.innerHTML=`<b style="color:${{sc}}">${{hn.label}}</b> <span style="color:#4a6a8a">${{hn.type}} · ${{hn.gate}}</span><br>
      <span style="color:#4a6a8a">${{hn.name}}</span><br>
      <span style="color:#5aabff">T = ${{hn.T}}</span>  <span style="color:${{sc}}">A = ${{hn.A}}</span>
      ${{hn.shared?'<br><span style="color:#fbbf24">⚡ Shared failure (synced)</span>':''}}`;
    tip.style.display='block';
    tip.style.left=Math.min(ev.clientX-r.left+14,cv.width-270)+'px';
    tip.style.top=Math.max(ev.clientY-r.top-60,0)+'px';
  }} else {{
    tip.style.display='none';
  }}
}});
window.addEventListener('mouseup',()=>{{
  drag=null;isPan=false;cv.classList.remove('pan');
}});
cv.addEventListener('wheel',ev=>{{
  ev.preventDefault();
  const r=cv.getBoundingClientRect();
  const cx=ev.clientX-r.left,cy=ev.clientY-r.top;
  const d=ev.deltaY<0?1.12:0.89;
  const ns=Math.max(0.05,Math.min(8,scale*d));
  panX=cx-(cx-panX)*(ns/scale);panY=cy-(cy-panY)*(ns/scale);scale=ns;
}},{{passive:false}});

// ── Search ──────────────────────────────────────────────────
function onSearch(q){{
  searchHL.clear();
  if(!q)return;
  const ql=q.toLowerCase();
  NODES.forEach(n=>{{
    if(n.label.toLowerCase().includes(ql)||n.name.toLowerCase().includes(ql)||n.type.toLowerCase().includes(ql))
      searchHL.add(n.id);
  }});
  // Also highlight ancestor path
  const base=new Set(searchHL);
  base.forEach(id=>{{
    let cur=id,seen=new Set();
    while(true){{
      const nd=NODES.find(x=>x.id===cur);if(!nd||!nd.parent||seen.has(cur))break;
      seen.add(cur);searchHL.add(nd.parent);cur=nd.parent;
    }}
  }});
  const first=[...base][0];
  if(first&&pos[first]){{panX=cv.width/2-pos[first].x*scale;panY=cv.height/3-pos[first].y*scale;}}
}}

function zoomIn(){{scale=Math.min(8,scale*1.2);}}
function zoomOut(){{scale=Math.max(0.05,scale/1.2);}}
function resetView(){{scale=1;panX=0;panY=60;searchHL.clear();layout();}}
function toggleSim(){{sim=!sim;document.getElementById('simBtn').textContent='⟳ Physics '+(sim?'ON':'OFF');}}

// ── Loop ──────────────────────────────────────────────────
function loop(){{simulate();draw();requestAnimationFrame(loop);}}
loop();
</script></body></html>"""

# ══════════════════════════════════════════════════════════════════════
# SIDEBAR — Node builder
# ══════════════════════════════════════════════════════════════════════
def sci_input(label, key_m, key_e, default_val=None):
    """Two-field scientific notation input. Returns float or None."""
    if default_val is not None and default_val > 0:
        exp = int(math.floor(math.log10(default_val)))
        man = round(default_val / (10 ** exp), 3)
    else:
        exp, man = -7, 1.0
    c1, c2 = st.columns([3, 2])
    m = c1.number_input(f"{label} ×", value=man, min_value=0.0, max_value=9.999,
                        step=0.001, format="%.3f", key=key_m, label_visibility="visible")
    e = c2.number_input("10^", value=exp, min_value=-20, max_value=0,
                        step=1, key=key_e, label_visibility="visible")
    if m > 0:
        return m * (10 ** e)
    return None

with st.sidebar:
    nodes = st.session_state.nodes
    hz_targets = st.session_state.hz_targets

    # Save status
    last_saved = st.session_state.get("_last_saved", "")
    if last_saved:
        st.markdown(f'<div class="save-ok">✓ Auto-saved · {last_saved}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="save-no">○ Not saved yet</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### ⚛ FTA Builder")

    # ── Step 1: Add Hazard ────────────────────────────────────
    with st.expander("① Add Hazard (HZ)", expanded=not nodes):
        hz_label = st.text_input("Label", value="HZ01", key="hz_lbl")
        hz_name  = st.text_input("Name", value="", key="hz_nm")
        hz_tgt   = sci_input("Target", "hz_m", "hz_e", 1e-7)
        if st.button("➕ Add Hazard", use_container_width=True):
            nid = next_id()
            nodes[nid] = {"id": nid, "label": hz_label, "name": hz_name,
                          "type": "HZ", "gate": "–", "parent": None, "achieved": None}
            hz_targets[nid] = hz_tgt or 1e-7
            save_to_file(); st.rerun()

    # ── Step 2–5: Add other nodes ─────────────────────────────
    ntype = st.selectbox("Node type", ["SF", "FF", "IF", "AND"], key="ntype")

    valid_parent_types = VALID_PARENTS.get(ntype, [])
    parent_options = {k: f"{v.get('label',k)} [{v['type']}]"
                      for k, v in nodes.items()
                      if v["type"] in valid_parent_types}

    if parent_options:
        with st.expander(f"② Add {ntype} node", expanded=False):
            n_label = st.text_input("Label", key="n_lbl")
            n_name  = st.text_input("Name",  key="n_nm")
            n_desc  = st.text_input("Desc",  key="n_dsc")
            n_par   = st.selectbox("Parent", list(parent_options.keys()),
                                   format_func=lambda k: parent_options[k], key="n_par")

            gate_default = "–" if ntype == "IF" else "OR"
            if ntype not in ("IF",):
                n_gate = st.selectbox("Gate (this→parent)", ["OR", "AND"], key="n_gate")
            else:
                n_gate = "–"
                st.info("IF nodes are leaves — no gate needed.")

            if st.button(f"➕ Add {ntype}", use_container_width=True):
                if not n_label.strip():
                    st.error("Label required")
                else:
                    nid = next_id()
                    nodes[nid] = {"id": nid, "label": n_label.strip(), "name": n_name,
                                  "desc": n_desc, "type": ntype, "gate": n_gate,
                                  "parent": n_par, "achieved": None}
                    save_to_file(); st.rerun()
    else:
        if nodes:
            st.caption(f"Add a {'/'.join(valid_parent_types)} node first to attach {ntype} to.")

    st.markdown("---")

    # ── Edit / Delete ─────────────────────────────────────────
    if nodes:
        with st.expander("✏️ Edit node", expanded=False):
            edit_opts = {k: f"{v.get('label',k)} [{v['type']}]" for k,v in nodes.items()}
            ek = st.selectbox("Select", list(edit_opts.keys()),
                              format_func=lambda k: edit_opts[k], key="ek")
            if ek and ek in nodes:
                en = nodes[ek]
                el = st.text_input("Label", value=en.get("label",""), key="el")
                ename = st.text_input("Name",  value=en.get("name",""),  key="ename")
                edesc = st.text_input("Desc",  value=en.get("desc",""),  key="edesc")
                if en["type"] == "HZ":
                    cur_tgt = hz_targets.get(ek, 1e-7)
                    new_tgt = sci_input("HZ Target", "etm", "ete", cur_tgt)
                if en["type"] not in ("HZ","IF"):
                    eg_opts = ["OR","AND"]
                    eg = st.selectbox("Gate", eg_opts,
                                      index=eg_opts.index(en.get("gate","OR")) if en.get("gate","OR") in eg_opts else 0,
                                      key="eg")
                else:
                    eg = en.get("gate","–")
                if st.button("💾 Save edit", use_container_width=True):
                    nodes[ek]["label"] = el
                    nodes[ek]["name"]  = ename
                    nodes[ek]["desc"]  = edesc
                    nodes[ek]["gate"]  = eg
                    if en["type"] == "HZ":
                        hz_targets[ek] = new_tgt or 1e-7
                    save_to_file(); st.success("Saved"); st.rerun()

        with st.expander("🗑 Delete node", expanded=False):
            del_opts = {k: f"{v.get('label',k)} [{v['type']}]" for k,v in nodes.items()}
            dk = st.selectbox("Node to delete", list(del_opts.keys()),
                              format_func=lambda k: del_opts[k], key="dk")
            if st.button("🗑 Delete + children", use_container_width=True):
                to_del = [dk] + descendants(nodes, dk)
                for d in to_del:
                    nodes.pop(d, None)
                    hz_targets.pop(d, None)
                save_to_file(); st.rerun()

    st.markdown("---")

    # ── File ops ──────────────────────────────────────────────
    with st.expander("💾 Save / Load", expanded=False):
        # Download
        st.markdown("**Download project as JSON**")
        if nodes:
            jstr = json.dumps(state_to_dict(), indent=2)
            st.download_button("⬇ Download JSON", data=jstr,
                               file_name="fta_project.json", mime="application/json",
                               use_container_width=True)
        else:
            st.caption("No data to download yet.")

        st.markdown("**Upload project JSON**")
        up = st.file_uploader("Upload JSON", type=["json"], key="up_json", label_visibility="collapsed")
        if up:
            try:
                d = json.load(up)
                dict_to_state(d)
                save_to_file()
                st.success("✓ Loaded!"); st.rerun()
            except Exception as ex:
                st.error(f"Load failed: {ex}")

    st.markdown("---")
    if st.button("⚠ Reset everything", use_container_width=True):
        st.session_state.nodes = {}
        st.session_state.hz_targets = {}
        st.session_state.nxt = 1
        if os.path.exists(SAVE_FILE):
            os.remove(SAVE_FILE)
        st.rerun()

# ══════════════════════════════════════════════════════════════════════
# MAIN AREA
# ══════════════════════════════════════════════════════════════════════
nodes      = st.session_state.nodes
hz_targets = st.session_state.hz_targets

# Compute both pipelines
alloc  = allocate(nodes, hz_targets)
rolled = rollup(nodes)
order  = bfs_order(nodes)

# Header
st.markdown("""<div class="app-header">
  <h1>⚛ FTA Risk Allocator v10</h1>
  <p>Reverse-engineer failure probability budgets · OR & AND gate logic · Shared failure sync · Auto-save</p>
</div>""", unsafe_allow_html=True)

# Quick stats
hz_list = [n for n in nodes.values() if n["type"]=="HZ"]
n_sf  = sum(1 for v in nodes.values() if v["type"]=="SF")
n_ff  = sum(1 for v in nodes.values() if v["type"] in ("FF","AND"))
n_if  = sum(1 for v in nodes.values() if v["type"]=="IF")
n_if_set = sum(1 for v in nodes.values() if v["type"]=="IF" and v.get("achieved") is not None)

if nodes:
    c1,c2,c3,c4,c5 = st.columns(5)
    def _mc(lbl, val, col):
        return f'<div style="background:#0c1220;border:1px solid #1e2d45;border-radius:6px;padding:8px 14px"><div style="font-size:0.6rem;color:#3a5a7a;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px">{lbl}</div><div style="font-family:JetBrains Mono,monospace;font-size:1rem;font-weight:700;color:{col}">{val}</div></div>'
    c1.markdown(_mc("Hazards", len(hz_list), "#fb923c"), unsafe_allow_html=True)
    c2.markdown(_mc("Sys Failures", n_sf, "#60a5fa"), unsafe_allow_html=True)
    c3.markdown(_mc("Flw/AND", n_ff, "#34d399"), unsafe_allow_html=True)
    c4.markdown(_mc("Init Events", n_if, "#c084fc"), unsafe_allow_html=True)
    c5.markdown(_mc("IF Values set", f"{n_if_set}/{n_if}", "#fbbf24"), unsafe_allow_html=True)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# Tabs
if not nodes:
    st.markdown('<div class="callout">👈 Start by adding a <b>Hazard (HZ)</b> node in the sidebar with its target failure rate, then add SF → FF → IF nodes below it.</div>', unsafe_allow_html=True)
else:
    tab_tree, tab_vals, tab_table, tab_export = st.tabs(["🌳 Tree", "✏️ Values", "📋 Table", "📥 Export"])

    # ── TAB 1: TREE VIZ ──────────────────────────────────────
    with tab_tree:
        components.html(build_canvas(nodes, alloc, rolled), height=680, scrolling=False)

    # ── TAB 2: VALUES ─────────────────────────────────────────
    with tab_vals:
        st.markdown('<div class="callout">'
            '<b style="color:#7ab8e8">How it works:</b><br>'
            '▶ <b>Allocated (T)</b> = auto-calculated top-down from HZ target. OR: T÷n per child · AND: T^(1/n) per child. <em>Never influenced by achieved values.</em><br>'
            '▶ <b>Achieved (A)</b> = value you enter. Rolled up bottom-up: OR: ΣA · AND: ΠA.<br>'
            '▶ <b>Shared failures</b>: nodes with the same Label get worst-case (max) value synced automatically across all hazards.'
            '</div>', unsafe_allow_html=True)

        sync_log = []  # collect changes for display

        for hz in hz_list:
            hid = hz["id"]
            hz_rolled = rolled.get(hid)
            hz_tgt    = hz_targets.get(hid, 1e-7)
            status_ok = hz_rolled is not None and hz_rolled <= hz_tgt
            status_bad= hz_rolled is not None and hz_rolled > hz_tgt
            scls = "ok" if status_ok else ("warn" if status_bad else "")
            sicon= "✅" if status_ok else ("❌" if status_bad else "⬜")
            st.markdown(
                f'<div class="callout {scls}">{sicon} <b>{hz.get("label","?")} — {hz.get("name","")}</b>'
                f'&emsp; T={fmt(hz_tgt)} &emsp; A={fmt(hz_rolled)}</div>',
                unsafe_allow_html=True)

            sub_ids = [i for i in order
                       if i != hid and i in nodes
                       and ancestor_hz(nodes, i) == hid]

            # Header row
            hc = st.columns([1.0, 1.8, 2.0, 0.6, 1.6, 2.2, 1.6])
            for h, t in zip(hc, ["Depth","Label","Name","Type","Allocated (T)","Achieved (enter)","Rolled-up (A)"]):
                h.markdown(f"<span style='font-size:0.6rem;color:#3a5a7a;text-transform:uppercase;letter-spacing:1px'>{t}</span>",
                           unsafe_allow_html=True)

            changed_nid = None
            changed_val = None

            for nid in sub_ids:
                if nid not in nodes: continue
                n   = nodes[nid]
                t   = n["type"]
                alc = alloc.get(nid)
                ach = n.get("achieved")
                rol = rolled.get(nid)
                d   = depth_of(nodes, nid)
                peers = nodes_with_label(nodes, n.get("label",""))
                is_shared = len(peers) > 1
                indent = "&ensp;" * (d * 2)

                # Status colour for rolled value
                rol_col = ("#4ade80" if (rol is not None and alc is not None and rol <= alc)
                           else "#f87171" if (rol is not None and alc is not None and rol > alc)
                           else "#3a5a7a")

                cols = st.columns([1.0, 1.8, 2.0, 0.6, 1.6, 2.2, 1.6])

                cols[0].markdown(f"<span style='color:#2a4a6a;font-size:0.7rem'>{'└─'*d if d else '●'}</span>", unsafe_allow_html=True)
                shared_tag = '<span class="st-shared">⚡shared</span>' if is_shared else ""
                t_color = TYPE_STR.get(t, "#7ab8e8")
                lbl_txt = n.get('label', nid)
                cols[1].markdown(f"{indent}<span style='font-family:JetBrains Mono,monospace;font-size:0.8rem;color:{t_color}'>{lbl_txt}</span>{shared_tag}", unsafe_allow_html=True)
                cols[2].markdown(f"<span style='font-size:0.77rem;color:#5a7a9a'>{n.get('name','')}</span>", unsafe_allow_html=True)
                cols[3].markdown(f"<span class='nb nb-{t}'>{t}</span>", unsafe_allow_html=True)
                cols[4].markdown(f"<span class='vbadge vb-alloc'>{fmt(alc)}</span>", unsafe_allow_html=True)

                with cols[5]:
                    ca, cb, cc = st.columns([2, 1.5, 0.8])
                    m_val = round(ach / (10 ** int(math.floor(math.log10(ach)))), 3) if (ach and ach > 0) else 1.0
                    e_val = int(math.floor(math.log10(ach))) if (ach and ach > 0) else -3
                    new_m = ca.number_input("m", value=m_val, min_value=0.0, max_value=9.999,
                                            step=0.001, format="%.3f",
                                            key=f"vm_{nid}", label_visibility="collapsed")
                    new_e = cb.number_input("e", value=e_val, min_value=-20, max_value=0,
                                            step=1, key=f"ve_{nid}", label_visibility="collapsed")
                    if cc.button("✕", key=f"vc_{nid}", help="Clear"):
                        nodes[nid]["achieved"] = None
                        save_to_file(); st.rerun()
                    else:
                        new_val = new_m * (10 ** new_e) if new_m > 0 else None
                        if new_val != ach:
                            changed_nid = nid
                            changed_val = new_val

                ach_cls = "vb-ach" + (" over" if (rol is not None and alc is not None and rol > alc) else "")
                badge_cls = ach_cls if rol is not None else "vb-none"
                cols[6].markdown(
                    f"<span class='vbadge {badge_cls}'>{fmt(rol)}</span>",
                    unsafe_allow_html=True)

            if changed_nid is not None and changed_val is not None:
                log = sync_shared(nodes, changed_nid, changed_val)
                sync_log.extend(log)
                save_to_file()
                st.rerun()

            st.markdown("<hr>", unsafe_allow_html=True)

        if sync_log:
            with st.expander(f"🔄 Shared sync — {len(sync_log)} node(s) updated", expanded=True):
                for nid, old, new in sync_log:
                    lbl = nodes.get(nid, {}).get("label", nid)
                    st.markdown(f"`{lbl}`: {fmt(old)} → {fmt(new)}")

    # ── TAB 3: TABLE ─────────────────────────────────────────
    with tab_table:
        search_t = st.text_input("🔍 Filter", placeholder="label / name / type…", key="tbl_srch")
        rows = []
        for nid in order:
            if nid not in nodes: continue
            n   = nodes[nid]
            alc = alloc.get(nid)
            rol = rolled.get(nid)
            par = nodes[n["parent"]]["label"] if n.get("parent") and n["parent"] in nodes else "–"
            if search_t:
                q = search_t.lower()
                if not (q in n.get("label","").lower() or q in n.get("name","").lower() or q in n["type"].lower()):
                    continue
            peers = nodes_with_label(nodes, n.get("label",""))
            rows.append({
                "Type": n["type"], "Label": n.get("label",nid), "Name": n.get("name",""),
                "Parent": par, "Gate": n.get("gate","–"),
                "Allocated T": fmt(alc), "Achieved A": fmt(rol),
                "Within Budget": ("✅" if (alc and rol and rol <= alc) else
                                  "❌" if (alc and rol and rol > alc) else "–"),
                "Shared": "⚡" if len(peers) > 1 else "",
            })

        if rows:
            import pandas as pd
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True,
                         column_config={
                             "Allocated T": st.column_config.TextColumn(help="Top-down budget from HZ target"),
                             "Achieved A":  st.column_config.TextColumn(help="Bottom-up rollup of entered values"),
                         })
        else:
            st.info("No nodes match the filter." if search_t else "No nodes yet.")

    # ── TAB 4: EXPORT ─────────────────────────────────────────
    with tab_export:
        col_json, col_xlsx = st.columns(2)
        with col_json:
            st.markdown("**JSON (full project)**")
            jdata = json.dumps(state_to_dict(), indent=2)
            st.download_button("⬇ Download JSON", data=jdata,
                               file_name="fta_project.json", mime="application/json",
                               use_container_width=True)
            st.caption("Reload this file to restore your exact session.")

        with col_xlsx:
            st.markdown("**Excel (.xlsx)**")
            def build_xlsx():
                wb = Workbook()
                ws = wb.active
                ws.title = "FTA_v10"
                hdrs = ["Type","Label","Name","Description","Parent","Gate",
                        "Allocated T","Achieved A","Within Budget","Shared Failure"]
                hfill = PatternFill("solid", fgColor="0c1e38")
                hfont = Font(name="Consolas", bold=True, color="7ab8e8")
                for ci, h in enumerate(hdrs, 1):
                    cell = ws.cell(1, ci, h)
                    cell.fill = hfill; cell.font = hfont
                    cell.alignment = Alignment(horizontal="center")

                for row_i, nid in enumerate(order, 2):
                    if nid not in nodes: continue
                    n   = nodes[nid]
                    alc = alloc.get(nid)
                    rol = rolled.get(nid)
                    par = nodes[n["parent"]]["label"] if n.get("parent") and n["parent"] in nodes else "–"
                    peers = nodes_with_label(nodes, n.get("label",""))
                    within = ("YES" if (alc and rol and rol <= alc) else
                              "NO"  if (alc and rol and rol > alc) else "–")
                    row = [n["type"], n.get("label",""), n.get("name",""), n.get("desc",""),
                           par, n.get("gate","–"), fmt(alc), fmt(rol), within,
                           "YES" if len(peers) > 1 else "NO"]
                    for ci, v in enumerate(row, 1):
                        ws.cell(row_i, ci, v)

                ws.column_dimensions["B"].width = 12
                ws.column_dimensions["C"].width = 28
                ws.column_dimensions["G"].width = 15
                ws.column_dimensions["H"].width = 15
                buf = io.BytesIO()
                wb.save(buf); buf.seek(0)
                return buf.getvalue()

            st.download_button("⬇ Download Excel", data=build_xlsx(),
                               file_name="fta_project.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)


